# procurement_to_excel_workers_final.py
import json
import time
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# === Configuration ===
BASE_URL = "http://localhost:8080/procurement/"
START_DATE = datetime(2022, 8, 1, 0, 0, 0)
END_DATE = datetime(2022, 4, 1, 0, 0, 0)
TIME_STEP = timedelta(minutes=15)

PRICE_CAP = 10.0
TIMEOUT = 30
MAX_RETRIES = 3
MAX_WORKERS = 8
PER_CALL_DELAY_SECONDS = 0.0  # if your API needs pacing, set small >0

# Single output file (stable name)
RANGE_TAG = (
    f"{START_DATE.strftime('%Y%m%d_%H%M')}_to_{END_DATE.strftime('%Y%m%d_%H%M')}"
)
OUTPUT_FILE = f"Procurement_{RANGE_TAG}.xlsx"


def make_session() -> requests.Session:
    retries = Retry(
        total=MAX_RETRIES,
        backoff_factor=0.3,
        status_forcelist=[500, 502, 503, 504],
        allowed_methods=frozenset(["GET"]),
    )
    adapter = HTTPAdapter(
        max_retries=retries, pool_connections=MAX_WORKERS, pool_maxsize=MAX_WORKERS
    )
    s = requests.Session()
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    return s


def flatten_record(data: dict, ts_str: str) -> dict:
    rec = dict(data)
    rec["Timestamp"] = ts_str

    iex = rec.pop("IEX_Data", {}) or {}
    rec["IEX_Pred_Price"] = iex.get("Pred_Price")
    rec["IEX_Qty_Pred"] = iex.get("Qty_Pred")

    for k in ("Must_Run", "Remaining_Plants"):
        if k in rec and isinstance(rec[k], (list, dict)):
            try:
                rec[k] = json.dumps(rec[k], separators=(",", ":"))
            except Exception:
                rec[k] = str(rec[k])
    return rec


def fetch_one(ts: datetime):
    """Return (ts_str, flat_or_None, err_or_None)."""
    session = make_session()
    ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
    try:
        resp = session.get(
            BASE_URL,
            params={"start_date": ts_str, "price_cap": PRICE_CAP},
            timeout=TIMEOUT,
        )
        resp.raise_for_status()
        data = resp.json()
        flat = flatten_record(data, ts_str)
        if PER_CALL_DELAY_SECONDS > 0:
            time.sleep(PER_CALL_DELAY_SECONDS)
        return ts_str, flat, None
    except Exception as e:
        return ts_str, None, str(e)


def build_ts_list(start: datetime, end: datetime, step: timedelta):
    out, cur = [], start
    while cur <= end:
        out.append(cur)
        cur += step
    return out


def write_excel(results: dict[str, dict]):
    if not results:
        print("No data fetched. Nothing to write.")
        return

    ordered_ts = sorted(results.keys())
    rows = [results[ts] for ts in ordered_ts]

    # Header union in first-seen order
    header = list(rows[0].keys())
    seen = set(header)
    for r in rows[1:]:
        for k in r.keys():
            if k not in seen:
                header.append(k)
                seen.add(k)

    wb = Workbook()
    ws = wb.active
    ws.title = "Procurement"

    ws.append(header)
    for rec in rows:
        ws.append([rec.get(k) for k in header])

    # Auto-fit columns (capped)
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max_len + 2, 60
        )

    wb.save(OUTPUT_FILE)
    print(f"✅ Data saved to '{OUTPUT_FILE}'")


def main():
    timestamps = build_ts_list(START_DATE, END_DATE, TIME_STEP)
    results: dict[str, dict] = {}
    errors: list[tuple[str, str]] = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(fetch_one, ts): ts for ts in timestamps}
        for fut in tqdm(
            as_completed(futures),
            total=len(futures),
            desc="Fetching procurement blocks",
        ):
            ts_str, flat, err = fut.result()
            if err:
                errors.append((ts_str, err))
            elif flat is not None:
                results[ts_str] = flat

    write_excel(results)

    if errors:
        print(f"\n⚠️  {len(errors)} blocks failed. Sample:")
        for ts_str, err in errors[:10]:
            print(f"  [{ts_str}] ❌ {err}")


if __name__ == "__main__":
    main()
